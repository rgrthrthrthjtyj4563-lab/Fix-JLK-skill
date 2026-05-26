#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def normalize_pct(value: object) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        return ""
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        number = float(text)
        if number <= 1:
            number *= 100
        return f"{number:.2f}%"
    except Exception:
        return f"{text}%"


def _clean_number(value: object) -> str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
        if number.is_integer():
            return str(int(number))
        return str(number)
    except Exception:
        return str(value).strip()


def _find_total_col(ws, question_row: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        value = ws.cell(question_row, col).value
        if str(value or "").strip() == "总计":
            return col
    return None


def parse_sheet(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    questions = []
    row = 2
    while row <= ws.max_row:
        q_text = ws.cell(row, 1).value
        if not q_text:
            row += 1
            continue
        count_row = row + 1
        pct_row = row + 2
        options = []
        total_col = _find_total_col(ws, row)
        total_value = ws.cell(count_row, total_col).value if total_col else None
        for col in range(3, ws.max_column + 1):
            raw = ws.cell(row, col).value
            if raw in (None, "", "总计"):
                continue
            raw_text = str(raw).strip()
            label = raw_text.split(".", 1)[0].strip()
            if label not in list("ABCDEF"):
                continue
            text = raw_text.split(".", 1)[1].strip() if "." in raw_text else raw_text
            if not text or text == label:
                continue
            options.append(
                {
                    "label": label,
                    "text": text,
                    "count": _clean_number(ws.cell(count_row, col).value),
                    "pct": normalize_pct(ws.cell(pct_row, col).value),
                }
            )
        questions.append(
            {
                "number": len(questions) + 1,
                "question": str(q_text).strip(),
                "options": options,
                "total": _clean_number(total_value) or None,
            }
        )
        row += 3
    return {
        "source_file": str(path),
        "sheet": ws.title,
        "question_count": len(questions),
        "questions": questions,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse questionnaire spreadsheet into normalized JSON.")
    parser.add_argument("input_file")
    parser.add_argument("-o", "--output", required=True)
    args = parser.parse_args()

    data = parse_sheet(Path(args.input_file))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
